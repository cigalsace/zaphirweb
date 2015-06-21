#!/usr/bin/env python
# -*- coding: utf-8 -*-
""" Module docstring
"""

__author__ = "Guillaume Ryckelynck"
__copyright__ = "Copyright 2015, Guillaume Ryckelynck"
__credits__ = ["Guillaume Ryckelynck"]
__license__ = "GPL"
__version__ = "0.1"
__maintainer__ = "Guillaume Ryckelynck"
__email__ = "guillaume.ryckelynck@region-alsace.org"
__status__ = "Developement"

import sys
import os
import json
import datetime
import xml.dom.minidom
import xml.etree.ElementTree
import uuid
import copy
import codecs
import re

try:
    import openpyxl
    openpyxl_actif = True
except ImportError:
    print "Error with openpyxl module"
    openpyxl_actif = False

def _get_json(filename=None):
    extension = os.path.splitext(filename)[1]
    if os.path.isfile(filename) and extension == '.json':
        with open(filename) as file:
            return json.load(file)

###########################################################################
# __DICT => XLSX
###########################################################################
def dict_to_xlsx(md_dict={}, xlsx_file=None, xlsx_template=None):
    """Convert data object to XLSX"""
    _root = os.path.dirname(__file__)
    # Get md_config.json
    _json_config = _get_json(os.path.join(_root, "json/md_config.json"))
    # Get md_codelist.json file
    _json_codelist = _get_json(os.path.join(_root, "json/md_codelist.json"))

    if xlsx_template is None or not os.path.isfile(xlsx_template):
        xlsx_template = os.path.join(root, _json_config['xlsx_template'])
    
    # convert data to XLSX and return file path
    if md_dict:
        #wb = openpyxl.load_workbook(xlsx_template, use_iterators=True)
        wb = openpyxl.load_workbook(xlsx_template)
        # ws = wb.get_active_sheet()
        named_ranges = wb.get_named_ranges()

        _put_xlsx_dict_value(wb, named_ranges, 'Info_FileIdentifier', md_dict, 'md_fileidentifier')
        _put_xlsx_dict_value(wb, named_ranges, 'Info_Language', md_dict, 'md_language', 'LanguageCode2', _json_codelist)
        _put_xlsx_dict_value(wb, named_ranges, 'Info_Characterset', md_dict, 'md_characterset', 'CharacterSetCode2', _json_codelist)
        _put_xlsx_dict_value(wb, named_ranges, 'Info_Hierarchylevel', md_dict, 'md_hierarchylevel', 'ScopeCode2', _json_codelist)
        _put_xlsx_dict_value(wb, named_ranges, 'Info_DateStamp', md_dict, 'md_datestamp')
        _put_xlsx_dict_value(wb, named_ranges, 'Info_Standardname', md_dict, 'md_standardname')
        _put_xlsx_dict_value(wb, named_ranges, 'Info_Standardversion', md_dict, 'md_standardversion')
        
        if 'md_contacts' in md_dict:
            for key, value in enumerate(md_dict['md_contacts']):
                _put_xlsx_dict_value(wb, named_ranges, 'Info_Contact'+str(key+1)+'_Name', value, 'name')
                _put_xlsx_dict_value(wb, named_ranges, 'Info_Contact'+str(key+1)+'_Position', value, 'position')
                _put_xlsx_dict_value(wb, named_ranges, 'Info_Contact'+str(key+1)+'_Organisation', value, 'organisation')
                _put_xlsx_dict_value(wb, named_ranges, 'Info_Contact'+str(key+1)+'_Tel', value, 'tel')
                _put_xlsx_dict_value(wb, named_ranges, 'Info_Contact'+str(key+1)+'_Email', value, 'email')
                _put_xlsx_dict_value(wb, named_ranges, 'Info_Contact'+str(key+1)+'_Address', value, 'address')
                _put_xlsx_dict_value(wb, named_ranges, 'Info_Contact'+str(key+1)+'_CP', value, 'cp')
                _put_xlsx_dict_value(wb, named_ranges, 'Info_Contact'+str(key+1)+'_City', value, 'city')
                _put_xlsx_dict_value(wb, named_ranges, 'Info_Contact'+str(key+1)+'_Role', value, 'role', 'RoleCode2', _json_codelist)

        _put_xlsx_dict_value(wb, named_ranges, 'Data_Title', md_dict, 'data_title')

        if 'data_dates' in md_dict:
            for key, value in enumerate(md_dict['data_dates']):
                if 'datetype' in value:
                    print value['datetype']
                    if value['datetype'] == 'creation':
                        _put_xlsx_dict_value(wb, named_ranges, 'data_datecreation', value, 'date')
                    elif value['datetype'] == 'publication':
                        _put_xlsx_dict_value(wb, named_ranges, 'data_datepublication', value, 'date')
                    elif value['datetype'] == 'revision':
                        _put_xlsx_dict_value(wb, named_ranges, 'data_daterevision', value, 'date')
                        
        """
        "data_identifiers": [
            {
                "code": "",
                "codespace": ""
            }
        ],
        """
        if 'data_identifiers' in md_dict:
            for key, value in enumerate(md_dict['data_identifiers']):
                _put_xlsx_dict_value(wb, named_ranges, 'Data_Identifier'+str(key+1)+'_Code', value, 'code')
                _put_xlsx_dict_value(wb, named_ranges, 'Data_Identifier'+str(key+1)+'_CodeSpace', value, 'codespace')
            
        
        # "data_abstract": "",
        _put_xlsx_dict_value(wb, named_ranges, 'Data_Abstract', md_dict, 'data_abstract')
        
        """
        "data_browsegraphics": [
            {
                "url": "",
                "description": "",
                "type": ""
            }
        ],
        """
        if 'data_browsegraphics' in md_dict:
            for key, value in enumerate(md_dict['data_browsegraphics']):
                _put_xlsx_dict_value(wb, named_ranges, 'Data_Browsegraphic'+str(key+1)+'_url', value, 'url')
                _put_xlsx_dict_value(wb, named_ranges, 'Data_Browsegraphic'+str(key+1)+'_description', value, 'description')
                _put_xlsx_dict_value(wb, named_ranges, 'Data_Browsegraphic'+str(key+1)+'_type', value, 'type')
                
       
        #"data_maintenancefrequencycode": "unknown",
        _put_xlsx_dict_value(wb, named_ranges, 'Data_MaintenanceFrequency', md_dict, 'data_maintenancefrequencycode', 'MaintenanceFrequencyCode2', _json_codelist)
        
        """
        "data_temporalextents": [
            {
                "start": "",
                "end": "",
                "description": ""
            }
        ],
        """
        if 'data_temporalextents' in md_dict:
            for key, value in enumerate(md_dict['data_temporalextents']):
                _put_xlsx_dict_value(wb, named_ranges, 'Data_TemporalExtent'+str(key+1)+'_Start', value, 'start')
                _put_xlsx_dict_value(wb, named_ranges, 'Data_TemporalExtent'+str(key+1)+'_End', value, 'end')
                _put_xlsx_dict_value(wb, named_ranges, 'Data_TemporalExtent'+str(key+1)+'_Description', value, 'description')
                
        #"data_languages": ["fre"],
        if 'data_languages' in md_dict:
            for key, value in enumerate(md_dict['data_languages']):
                _put_xlsx_list_value(wb, named_ranges, 'Data_Language'+str(key+1), value, 'LanguageCode2', _json_codelist)
        
        #"data_topiccategories": [""],
        if 'data_topiccategories' in md_dict:
            for key, value in enumerate(md_dict['data_topiccategories']):
                _put_xlsx_list_value(wb, named_ranges, 'Data_TopiCcategory'+str(key+1), value, 'TopicCategoryCode2', _json_codelist)

        """
        "data_inspirekeywords": [
            {
                "keyword": "",
                "type": "",
                "thesaurus_name": "GEMET - INSPIRE themes, version 1.0",
                "thesaurus_dates": [
                    {
                        "type": "publication",
                        "date": "2008-06-01"
                    }
                ]
            }
        ],
        """
        if 'data_inspirekeywords' in md_dict:
            for key, value in enumerate(md_dict['data_inspirekeywords']):
                _put_xlsx_dict_value(wb, named_ranges, 'Data_InspireKeyword'+str(key+1)+'_keyword', value, 'keyword', 'InspireThemeCode_en2', _json_codelist)
                _put_xlsx_dict_value(wb, named_ranges, 'Data_InspireKeyword'+str(key+1)+'_Type', value, 'type')
                _put_xlsx_dict_value(wb, named_ranges, 'Data_InspireKeyword'+str(key+1)+'_ThesaurusName', value, 'thesaurus_name')
                
                if 'thesaurus_dates' in md_dict['data_inspirekeywords']:
                    for key2, value2 in enumerate(md_dict['data_inspirekeywords']['thesaurus_dates']):
                        if 'type' in value2:
                            if value2['type'] == 'creation':
                                _put_xlsx_dict_value(wb, named_ranges, 'Data_InspireKeyword'+str(key+1)+'_ThesaurusDateCreation', value2, 'date')
                            elif value2['type'] == 'publication':
                                _put_xlsx_dict_value(wb, named_ranges, 'Data_InspireKeyword'+str(key+1)+'_ThesaurusDatePublication', value2, 'date')
                            elif value2['type'] == 'revision':
                                _put_xlsx_dict_value(wb, named_ranges, 'Data_InspireKeyword'+str(key+1)+'_ThesaurusDateRevision', value2, 'date')

        """
        "data_keywords": [
            {
                "keyword": "données ouvertes",
                "type": "theme",
                "thesaurus_name": "",
                "thesaurus_dates": [
                    {
                        "type": "",
                        "date": ""
                    }
                ]
            },
            {
                "keyword": "Géoportail",
                "type": "theme",
                "thesaurus_name": "",
                "thesaurus_dates": [
                    {
                        "type": "",
                        "date": ""
                    }
                ]
            },
            {
                "keyword": "",
                "type": "",
                "thesaurus_name": "",
                "thesaurus_dates": [
                    {
                        "type": "",
                        "date": ""
                    }
                ]
            }
        ],
        """
        if 'data_keywords' in md_dict:
            for key, value in enumerate(md_dict['data_keywords']):
                _put_xlsx_dict_value(wb, named_ranges, 'Data_Keyword'+str(key+1)+'_keyword', value, 'keyword')
                _put_xlsx_dict_value(wb, named_ranges, 'Data_Keyword'+str(key+1)+'_Type', value, 'type')
                _put_xlsx_dict_value(wb, named_ranges, 'Data_Keyword'+str(key+1)+'_ThesaurusName', value, 'thesaurus_name')
                
                if 'thesaurus_dates' in md_dict['data_keywords']:
                    for key2, value2 in enumerate(md_dict['data_keywords']['thesaurus_dates']):
                        if 'type' in value2:
                            if value2['type'] == 'creation':
                                _put_xlsx_dict_value(wb, named_ranges, 'Data_Keyword'+str(key+1)+'_ThesaurusDateCreation', value2, 'date')
                            elif value2['type'] == 'publication':
                                _put_xlsx_dict_value(wb, named_ranges, 'Data_Keyword'+str(key+1)+'_ThesaurusDatePublication', value2, 'date')
                            elif value2['type'] == 'revision':
                                _put_xlsx_dict_value(wb, named_ranges, 'Data_Keyword'+str(key+1)+'_ThesaurusDateRevision', value2, 'date')
        
        """
        "data_contacts": [
            {
                "name": "",
                "position": "",
                "organisation": "",
                "tel": "",
                "email": "",
                "address": "",
                "cp": "",
                "city": "",
                "logo_url": "",
                "logo_text": "",
                "role": ""
            }
        ],
        """
        if 'data_contacts' in md_dict:
            for key, value in enumerate(md_dict['data_contacts']):
                _put_xlsx_dict_value(wb, named_ranges, 'Data_Contact'+str(key+1)+'_Name', value, 'name')
                _put_xlsx_dict_value(wb, named_ranges, 'Data_Contact'+str(key+1)+'_Position', value, 'position')
                _put_xlsx_dict_value(wb, named_ranges, 'Data_Contact'+str(key+1)+'_Organisation', value, 'organisation')
                _put_xlsx_dict_value(wb, named_ranges, 'Data_Contact'+str(key+1)+'_Tel', value, 'tel')
                _put_xlsx_dict_value(wb, named_ranges, 'Data_Contact'+str(key+1)+'_Email', value, 'email')
                _put_xlsx_dict_value(wb, named_ranges, 'Data_Contact'+str(key+1)+'_Address', value, 'address')
                _put_xlsx_dict_value(wb, named_ranges, 'Data_Contact'+str(key+1)+'_CP', value, 'cp')
                _put_xlsx_dict_value(wb, named_ranges, 'Data_Contact'+str(key+1)+'_City', value, 'city')
                _put_xlsx_dict_value(wb, named_ranges, 'Data_Contact'+str(key+1)+'_Role', value, 'role', 'RoleCode2', _json_codelist)
                
        """
        "data_geographicextents": [
            {
                "name": "",
                "xmin": "",
                "xmax": "",
                "ymin": "",
                "ymax": ""
            }
        ],
        """
        if 'data_geographicextents' in md_dict:
            for key, value in enumerate(md_dict['data_geographicextents']):
                _put_xlsx_dict_value(wb, named_ranges, 'Data_GeographicExtent'+str(key+1)+'_Name', value, 'name')
                _put_xlsx_dict_value(wb, named_ranges, 'Data_GeographicExtent'+str(key+1)+'_Xmin', value, 'xmin')
                _put_xlsx_dict_value(wb, named_ranges, 'Data_GeographicExtent'+str(key+1)+'_Xmax', value, 'xmax')
                _put_xlsx_dict_value(wb, named_ranges, 'Data_GeographicExtent'+str(key+1)+'_Ymin', value, 'ymin')
                _put_xlsx_dict_value(wb, named_ranges, 'Data_GeographicExtent'+str(key+1)+'_Ymax', value, 'ymax')

        """
        "data_referencesystems": [
            {
                "code": "",
                "codespace": ""
            }
        ],
        """
        if 'data_referencesystems' in md_dict:
            for key, value in enumerate(md_dict['data_referencesystems']):
                _put_xlsx_dict_value(wb, named_ranges, 'Data_ReferenceSystem'+str(key+1)+'_Code', value, 'code')
                _put_xlsx_dict_value(wb, named_ranges, 'Data_ReferenceSystem'+str(key+1)+'_CodeSpace', value, 'codespace')
    
        #"data_presentationform": "",
        #TODO
        
        #"data_spatialrepresentationtype": "",
        _put_xlsx_dict_value(wb, named_ranges, 'Data_SpatialRepresentationType', md_dict, 'data_spatialrepresentationtype', 'SpatialRepresentationTypeCode2', _json_codelist)
        
        #"data_scaledenominator": "",
        _put_xlsx_dict_value(wb, named_ranges, 'Data_ScaleDenominator', md_dict, 'data_scaledenominator')

        #"data_scaledistance": "",
        _put_xlsx_dict_value(wb, named_ranges, 'Data_ScaleDistance', md_dict, 'data_scaledistance')

        #"data_dq_level": "dataset",
        _put_xlsx_dict_value(wb, named_ranges, 'Data_DQ_Level', md_dict, 'data_dq_level', 'ScopeCode2', _json_codelist)
        
        #"data_li_statement": "",
        _put_xlsx_dict_value(wb, named_ranges, 'Data_LI_Statement', md_dict, 'data_li_statement')

        #"data_characterset": "utf8",
        _put_xlsx_dict_value(wb, named_ranges, 'Data_Characterset', md_dict, 'data_characterset', 'CharacterSetCode2', _json_codelist)
        
        """
        "data_distformats": [
            {
                "name": "",
                "version": "",
                "specification": ""
            }
        ],
        """
        if 'data_distformats' in md_dict:
            for key, value in enumerate(md_dict['data_distformats']):
                _put_xlsx_dict_value(wb, named_ranges, 'Data_Distformat'+str(key+1)+'_Name', value, 'name')
                _put_xlsx_dict_value(wb, named_ranges, 'Data_Distformat'+str(key+1)+'_Version', value, 'version')
                _put_xlsx_dict_value(wb, named_ranges, 'Data_Distformat'+str(key+1)+'_Specification', value, 'specification')

    
        # "data_uselimitations": ["ul1", "ul2"],
        if 'data_uselimitations' in md_dict:
            for key, value in enumerate(md_dict['data_uselimitations']):
                _put_xlsx_list_value(wb, named_ranges, 'Data_UseLimitation'+str(key+1), value)
        
        # "data_legal_uselimitations": ["Licence Etalab 2011 (http://ddata.over-blog.com/xxxyyy/4/37/99/26/licence/Licence-Ouverte-Open-Licence.pdf)", "l_ul2"],
        if 'data_legal_uselimitations' in md_dict:
            for key, value in enumerate(md_dict['data_legal_uselimitations']):
                _put_xlsx_list_value(wb, named_ranges, 'Data_Legal_UseLimitation'+str(key+1), value)
        
        # "data_legal_useconstraints": ["license", "otherRestrictions", "otherRestrictions"],
        if 'data_legal_useconstraints' in md_dict:
            for key, value in enumerate(md_dict['data_legal_useconstraints']):
                _put_xlsx_list_value(wb, named_ranges, 'Data_Legal_Useconstraint'+str(key+1), value, 'RestrictionCode2', _json_codelist)
        
        # "data_legal_accessconstraints": ["copyright", "otherRestrictions"],
        if 'data_legal_accessconstraints' in md_dict:
            for key, value in enumerate(md_dict['data_legal_accessconstraints']):
                _put_xlsx_list_value(wb, named_ranges, 'Data_Legal_Accessconstraint'+str(key+1), value, 'RestrictionCode2', _json_codelist)
        
        # "data_legal_otherconstraints": ["Pas de restriction d'accès public", "other2", "other3", "other4"],
        if 'data_legal_otherconstraints' in md_dict:
            for key, value in enumerate(md_dict['data_legal_otherconstraints']):
                _put_xlsx_list_value(wb, named_ranges, 'Data_Legal_Otherconstraint'+str(key+1), value, 'InspireRestrictionCode2', _json_codelist)
        
        # "data_security_uselimitations": ["s_ul1", "s_ul2"],
        if 'data_security_uselimitations' in md_dict:
            for key, value in enumerate(md_dict['data_security_uselimitations']):
                _put_xlsx_list_value(wb, named_ranges, 'Data_Security_Uselimitation'+str(key+1), value)
        
        # "data_security_classification": "unclassified",
        _put_xlsx_dict_value(wb, named_ranges, 'Data_Security_Classification', md_dict, 'data_security_classification', 'ClassificationCode2', _json_codelist)
        
        """
        "data_linkages": [
            {
                "name": "",
                "description": "",
                "url": "",
                "protocol": ""
            }
        ],
        """
        if 'data_linkages' in md_dict:
            for key, value in enumerate(md_dict['data_linkages']):
                _put_xlsx_dict_value(wb, named_ranges, 'Data_Linkage'+str(key+1)+'_Name', value, 'name')
                _put_xlsx_dict_value(wb, named_ranges, 'Data_Linkage'+str(key+1)+'_Description', value, 'description')
                _put_xlsx_dict_value(wb, named_ranges, 'Data_Linkage'+str(key+1)+'_Url', value, 'url')
                _put_xlsx_dict_value(wb, named_ranges, 'Data_Linkage'+str(key+1)+'_Protocol', value, 'protocol')

        """
        "data_dq_inspireconformities": [
            {
                "specification": "",
                "explaination": "",
                "pass": "",
                "dates": [
                    {
                        "type": "",
                        "date": ""
                    }
                ]
            }
        ]
        """
        if 'data_dq_inspireconformities' in md_dict:
            for key, value in enumerate(md_dict['data_dq_inspireconformities']):
                _put_xlsx_dict_value(wb, named_ranges, 'Data_DQ_InspireConformity'+str(key+1)+'_Specification', value, 'specification', 'InspireSpecificationCode2', _json_codelist)
                _put_xlsx_dict_value(wb, named_ranges, 'Data_DQ_InspireConformity'+str(key+1)+'_Explaination', value, 'explaination')
                _put_xlsx_dict_value(wb, named_ranges, 'Data_DQ_InspireConformity'+str(key+1)+'_Pass', value, 'pass')
                
                # Date des spécifications Inspire non gérées
                '''
                if 'dates' in md_dict['data_dq_inspireconformities']:
                    for key2, value2 in enumerate(md_dict['data_dq_inspireconformities']['dates']):
                        if 'type' in value2:
                            if value2['type'] == 'creation':
                                _put_xlsx_dict_value(wb, named_ranges, 'Data_Keyword'+str(key+1)+'_ThesaurusDateCreation', value2, 'date')
                            elif value2['type'] == 'publication':
                                _put_xlsx_dict_value(wb, named_ranges, 'Data_Keyword'+str(key+1)+'_ThesaurusDatePublication', value2, 'date')
                            elif value2['type'] == 'revision':
                                _put_xlsx_dict_value(wb, named_ranges, 'Data_Keyword'+str(key+1)+'_ThesaurusDateRevision', value2, 'date')
                '''
        
        """
        "data_dq_conformities": [
            {
                "specification": "",
                "explaination": "",
                "pass": "",
                "dates": [
                    {
                        "type": "",
                        "date": ""
                    }
                ]
            }
        ]
        """
        if 'data_dq_conformities' in md_dict:
            for key, value in enumerate(md_dict['data_dq_conformities']):
                _put_xlsx_dict_value(wb, named_ranges, 'Data_DQ_Conformity'+str(key+1)+'_Specification', value, 'specification', 'InspireSpecificationCode2', _json_codelist)
                _put_xlsx_dict_value(wb, named_ranges, 'Data_DQ_Conformity'+str(key+1)+'_Explaination', value, 'explaination')
                _put_xlsx_dict_value(wb, named_ranges, 'Data_DQ_Conformity'+str(key+1)+'_Pass', value, 'pass')
                
                if 'dates' in md_dict['data_dq_conformities']:
                    for key2, value2 in enumerate(md_dict['data_dq_conformities']['dates']):
                        if 'type' in value2:
                            if value2['type'] == 'creation':
                                _put_xlsx_dict_value(wb, named_ranges, 'Data_DQ_Conformity'+str(key+1)+'_DateCreation', value2, 'date')
                            elif value2['type'] == 'publication':
                                _put_xlsx_dict_value(wb, named_ranges, 'Data_DQ_Conformity'+str(key+1)+'_DatePublication', value2, 'date')
                            elif value2['type'] == 'revision':
                                _put_xlsx_dict_value(wb, named_ranges, 'Data_DQ_Conformity'+str(key+1)+'_DateRevision', value2, 'date')


        '''
        for data_name, data_value in md_dict.items():
            if data_name in md_xlsx:
                cell_info = md_xlsx[data_name]
                if type(data_value) is list:
                    for k, v in enumerate(data_value):
                        cell_name = cell_info['name'].replace('%', str(k+1))
                        _put_xlsx_dict_value(wb, named_ranges, cell_name, v)
                elif type(data_value) is dict:
                    pass

                else:
                    _put_xlsx_dict_value(wb, named_ranges, cell_info['name'], data_value)

        obj_list = []
        for data_name, cell_info in md_xlsx.items():
            if 'many' in cell_info and cell_info['many'] is not None:
                split = data_name.split(':')
                is_object = len(split) > 1
                if is_object:
                    if split[0] in md_dict:
                        pass
                else:
                    for id, item in enumerate(md_dict[split[0]]):
                        cell_info['name'] = cell_info['name'].replace('%', str(id))
                        _put_xlsx_dict_value(wb, named_ranges, cell_info['name'], item)

            else:    
                if data_name in md_dict:
                    _put_xlsx_dict_value(wb, named_ranges, cell_info['name'], md_dict[data_name])
                if 'many' in cell_info and cell_info['many']:
                    split = data_name.split(':')
                    is_object = len(split) > 1
                    if is_object:
                        if cell_info['type'] != 'date':
                            tmp = 'tmp_'+split[0]
                            if split[0] not in obj_list: obj_list.append(split[0])
                            if not tmp in md_dict: md_dict[tmp] = {}
                            for i in range(1, 10):
                                if not i in md_dict[tmp]: md_dict[tmp][i] = {}
                                cell_info_new = copy.deepcopy(cell_info)
                                cell_info_new['name'] = cell_info['name'].replace('%', str(i))
                                cell_value = _get_xls_value(wb, named_ranges, cell_info_new)
                                if cell_value:
                                    md_dict[tmp][i][split[1]] = cell_value
                        elif len(split) == 2:
                            # Gestion des dates simples
                            cell_value = _get_xls_value(wb, named_ranges, cell_info)
                            date = {'date': cell_value, 'type': split[1]}
                            if split[0] not in md_dict: md_dict[split[0]] = []
                            md_dict[split[0]].append(date)
                        elif len(split) == 3:
                            # Gestion des dates des thesaurus
                            tmp = 'tmp_'+split[0]
                            if split[0] not in obj_list: obj_list.append(split[0])
                            if not tmp in md_dict: md_dict[tmp] = {}
                            for i in range(1, 10):
                                if not i in md_dict[tmp]: md_dict[tmp][i] = {}
                                if not split[1] in md_dict[tmp][i]: md_dict[tmp][i][split[1]] = []
                                cell_info_new = copy.deepcopy(cell_info)
                                cell_info_new['name'] = cell_info['name'].replace('%', str(i))
                                cell_value = _get_xls_value(wb, named_ranges, cell_info_new)
                                if cell_value:
                                    md_dict[tmp][i][split[1]].append({'date': cell_value, 'type': split[2]})
                    else:
                        if not data_name in md_dict: md_dict[data_name] = []
                        for i in range(1, 10):
                            cell_info_new = copy.deepcopy(cell_info)
                            cell_info_new['name'] = cell_info['name'].replace('%', str(i))
                            cell_value = _get_xls_value(wb, named_ranges, cell_info_new)
                            if cell_value:
                                md_dict[data_name].append(cell_value)
                else:
                    _put_xlsx_dict_value(wb, named_ranges, cell_info['name'], md_dict[data_name])
                    cell_value = _get_xls_value(wb, named_ranges, cell_info)
                    if cell_value:
                        md_dict[cell_name] = cell_value

            '''

        #xlsx_file = os.path.join(root, _json_config['path_out'], 'test2.xlsx')
        wb.save(xlsx_file)

        #_get_codelist(wb, named_ranges, 'MD_LanguageCode')

"""
def _get_codelist(wb=None, named_ranges=None, codelist=None):
    cell = wb.get_named_range(codelist)
    print cell
"""

"""
def _put_xlsx_value(wb=None, named_ranges=None, cell_name=None, data_name=None, codelist=None):
    if data_name in md_dict:
        cell_names = [cell_name, cell_name.lower(), cell_name.upper()]
        for c_name in cell_names:
            cell = wb.get_named_range(c_name)
            if cell:
                ws = cell.destinations[0][0]
                #ws.show_gridlines = False
                c = ws.cell(cell.destinations[0][1])
                '''
                for row in ws.iter_rows(cell.destinations[0][1]):
                    for cell in row:
                        c = cell
                '''
                if codelist is not None and codelist in md_codelist:
                    inv_codelist = {v: k for k, v in md_codelist[codelist].items()}
                    if md_dict[data_name] in inv_codelist:
                        cell_value = inv_codelist[md_dict[data_name]]
                else:
                    cell_value = md_dict[data_name]
                c.value = cell_value
                return True
            else:
                return False
"""


def _put_xlsx_dict_value(wb=None, named_ranges=None, cell_name=None, dict_name={}, data_name=None, codelist_name=None, codelist={}):
    response = False
    if data_name and data_name in dict_name:
        cell_names = [cell_name, cell_name.lower(), cell_name.upper()]
        for c_name in cell_names:
            cell = wb.get_named_range(c_name)
            if cell: #c_name existe dans XLSX => récuppérer les coordonnées de cette cellule
                ws = cell.destinations[0][0]
                c = ws.cell(cell.destinations[0][1])
                if dict_name[data_name] and codelist_name is not None and codelist_name in codelist:
                    cell_value = codelist[codelist_name][dict_name[data_name]]
                else:
                    cell_value = dict_name[data_name]
                c.value = cell_value
                response = cell_value
    return response

def _put_xlsx_list_value(wb=None, named_ranges=None, cell_name=None, data_value=None, codelist_name=None, codelist={}):
    response = False
    if data_value:
        cell_names = [cell_name, cell_name.lower(), cell_name.upper()]
        for c_name in cell_names:
            cell = wb.get_named_range(c_name)
            if cell: #c_name existe dans XLSX => récuppérer les coordonnées de cette cellule
                ws = cell.destinations[0][0]
                c = ws.cell(cell.destinations[0][1])
                if codelist_name is not None and codelist_name in codelist:
                    cell_value = codelist[codelist_name][data_value]
                else:
                    cell_value = data_value
                c.value = cell_value
                response = cell_value
    return response

"""
def _put_xlsx_values(wb, named_ranges, cell_name, cell_values):
    response = False
    for cell_key, cell_value in enumerate(cell_values):
        c_name = cell_name.replace('%', str(cell_key+1))
        cell_names = [c_name, c_name.lower(), c_name.upper()]
        for c_name in cell_names:
            cell = wb.get_named_range(c_name)
            if cell:
                ws = cell.destinations[0][0]
                c = ws.cell(cell.destinations[0][1])
                if codelist is not None and codelist in md_codelist:
                    cell_value = codelist[md_dict[dict_name]]
                else:
                    cell_value = md_dict[dict_name]
                c.value = cell_value
                response = cell_value
    return response
"""

if __name__ == "__main__":
    _root = os.path.dirname(__file__)
    md_dict = _get_json(os.path.join(_root, "json/md_data.json"))       
    dict_to_xlsx(md_dict, 'test/xlsx/test.xlsx', 'test/xlsx/MODELE-VIDE_FicheMD-CIGAL_simple_150428.xlsx')

